import json
from jira import JIRA
import keyring
from keyrings.alt import Windows
keyring.set_keyring(Windows.RegistryKeyring())
import tkinter
from tkinter import Label
from tkinter import Button
from tkinter import Entry
from tkinter import Checkbutton
from tkinter import OptionMenu
from tkinter import StringVar
from tkinter import Listbox
from tkinter import Scrollbar
from tkinter import messagebox
from tkinter import N, S, E, W
from tkinter import VERTICAL
import openpyxl
import webbrowser
from openpyxl import Workbook, load_workbook
from threading import Thread

'''Below module allows us to interact with Windows files.'''
import os

'''below 3 lines allows script to check the directory where it is executed, so it knows where to crete the excel file. I copied the whole block from stack overflow'''
abspath = os.path.abspath(__file__)
current_directory = os.path.dirname(abspath)
os.chdir(current_directory)

'''Below lines set values for fixed variables. They are used in several functions and are set here for more convenient update in future if needed.'''
destination_excel_file_name = "Raw jira data.xlsx"
maximum_results = 1000

'''Below lines are placeholder names. They will be set to needed values by functions further. Class usage allow to get rid of declaring them here, but classes were disregarded for code simplicity.'''
login_done = 0
jira = ""

def get_files_in_script_directory():
    '''Get file names in directory'''
    file_names = []
    for root, dirs, files in os.walk(current_directory):
        for filename in files:
            file_names.append(filename)
    return file_names


def insert_text(text):
    text_box.insert('end', text)
    text_box.see("end")

def insert_log(project, number_of_issues):
    text_box.insert('end', str(project) + " : " + str(number_of_issues) + " issues where imported")
    text_box.see("end")


def open_raw_jira_data_excel():
    os.startfile(current_directory + "\\" + destination_excel_file_name, 'open') 
    

def open_apps_folder():
    os.startfile(current_directory + "\\", 'open') 
    

def open_web_link(event):
    webbrowser.open_new(r"https://www.google.com/")

def perform_login():
    '''Log's in if it was not already done on current App usage.'''
    global login_done
    global jira
    if login_done == 0:
        try:
            insert_text("Performing login")
            first_line_info = login_field_entry_var.get()
            second_line_info = keyring.get_password("jira", first_line_info)
            jira = JIRA("JIRA API URL GOES HERE", basic_auth=(first_line_info, second_line_info)) # https://api-jira.COMPANY_NAME.com/
            insert_text("Login succesful")
            login_done = 1
        except:
            insert_text("Login failed. Please, make sure you saved your credentials.")


def create_login_file_if_it_was_not_there():
    '''Creates login.txt if it was not found in current directory'''
    if "login.txt" not in get_files_in_script_directory():
        with open("login.txt", 'w') as f:
            f.write("Please, input login")


def get_search_settings():
    """Open search_settings.txt and and get search settings from there"""
    search_settings = ""
    with open("search_settings.txt", 'r') as f:
        search_settings = f.read()
    return search_settings

def create_search_settings_file_if_it_was_not_there():
    '''Creates search_settings.txt if it was not found in current directory'''
    if "search_settings.txt" not in get_files_in_script_directory():
        with open("search_settings.txt", 'w') as f:
            f.write('''project = PROJECT_NAME AND due >= 2018-01-01 AND due <= 2019-12-31''')

def get_needed_fields():
    """Open needed_fields.txt and and get a list of fields from there"""
    needed_fields = []
    with open("needed_fields.txt", 'r') as f:
        for field in f:
            needed_fields.append(field.replace("\n", ""))
    return needed_fields

def create_needed_fields_file_if_it_was_not_there():
    '''Creates needed_fields.txt if it was not found in current directory'''
    if "needed_fields.txt" not in get_files_in_script_directory():
        with open("needed_fields.txt", 'w') as f:
            for i in range(3):
                f.write('''customfield_00000\n''')

def save_login_and_password():
    '''Saves provided by user login to text file and password to windows registry with keyrings.alt module'''
    first_line_info = login_field_entry_var.get()
    second_line_info = password_field_entry_var.get()
    with open("login.txt", 'w') as f:
            f.write(first_line_info)
    keyring.set_password("jira", first_line_info, second_line_info)


def get_login_from_file():
    '''Get login saved in login.txt file'''
    first_line_info = ""
    with open("login.txt", 'r') as f:
        first_line_info = f.read()
    return first_line_info


def create_excel_file_if_it_was_not_there():
    '''Creates an excel file where jira search result will be stored if it was not found in current directory'''
    if destination_excel_file_name not in get_files_in_script_directory():
        wb = Workbook()
        ws = wb.active
        wb.save(destination_excel_file_name)


def find_the_row_of_the_next_empty_cell(wb, ws):
    '''Scan cells of destination excel file to find next empty cell'''
    row_to_check = 1
    cell_to_paste_link_to = ws['A' + str(row_to_check)]
    while cell_to_paste_link_to.value != None:
        row_to_check = row_to_check + 1
        cell_to_paste_link_to = ws['A' + str(row_to_check)]
    return row_to_check

def clear_existing_data():
    '''Delete existing jira data sheet and creates an empty one'''
    wb = load_workbook(destination_excel_file_name)
    ws = wb.active
    wb.remove(ws)
    wb.create_sheet("Raw data")
    ws = wb.active
    wb.save(destination_excel_file_name)
    wb = load_workbook(destination_excel_file_name)
    ws = wb.active
    return wb, ws


def write_data_to_excel(wb, ws):
    """Write jira data extracted with saved search options to excel"""
    search_settings = get_search_settings()
    needed_fileds = get_needed_fields()
    issues = jira.search_issues(search_settings, maxResults = maximum_results, fields = needed_fileds)
    current_row = find_the_row_of_the_next_empty_cell(wb, ws)
    for issue in issues:
        column_index = 1
        ws.cell(row=current_row, column=column_index, value=str(issue.key))
        column_index +=1
        ws.cell(row=current_row, column=column_index, value=int(str(issue.id)))
        column_index +=1
        for field in needed_fileds:
            try:
                ws.cell(row=current_row, column=column_index, value=str(issue.raw['fields'][field]))
            except:
                ws.cell(row=current_row, column=column_index, value="N/A")
                print("Was not able to extract " + field)
            column_index +=1
        current_row += 1


def write_combined_data_to_excel():
    def slow_magic():
        perform_login()
        insert_text("Starting gathering data from Jira...")
        insert_text("This can take several minutes...")
        wb, ws = clear_existing_data()
        insert_text("Extracting data...")
        write_data_to_excel(wb, ws)
        wb.save(destination_excel_file_name)
        insert_text("Import done.")
        open_raw_jira_data_excel()
    t = Thread(target=slow_magic)
    t.start()



def get_all_fields_to_json():
    perform_login()
    search_settings = get_search_settings()
    issues = jira.search_issues(search_settings, maxResults = 1) # search option example example: project = PROJECT_NAME AND due >= 2018-01-01 AND due <= 2019-12-31
    with open("project_all_fields.txt", "w") as project_all_fields_file:
        json.dump(issues[0].raw, project_all_fields_file)
    insert_text("All fields JSON saved to project_all_fields.txt.")


#********************file creation***********************
create_excel_file_if_it_was_not_there()
create_login_file_if_it_was_not_there()
create_search_settings_file_if_it_was_not_there()
create_needed_fields_file_if_it_was_not_there()

#********************interface defined here*******************
main_window_of_gui = tkinter.Tk()
main_window_of_gui.title("Jira scraper v0001")
main_window_of_gui.wm_attributes("-topmost", 1)

login_field_entry_var = StringVar()
password_field_entry_var = StringVar()

login_field_entry = Entry(main_window_of_gui, width = 20, textvariable = login_field_entry_var)
login_field_entry.grid(row = 0, column = 1)
login_field_entry_var.set(get_login_from_file())

password_field_entry = Entry(main_window_of_gui, show="*", width = 15, textvariable = password_field_entry_var)
password_field_entry.grid(row = 0, column = 2)
password_field_entry_var.set("password")

empty_column = Label(main_window_of_gui, text = "   ")
empty_column.grid(row = 0, column = 4, sticky=tkinter.W)

save_login_button = Button(main_window_of_gui, text = "Save", command = save_login_and_password)
save_login_button.grid(row = 0, column = 0)

extract_selected_fields_to_excel = Button(main_window_of_gui, text ="Import Data", width = 15, height = 3, command = write_combined_data_to_excel)
extract_selected_fields_to_excel.grid(row = 1, column = 2, rowspan = 3, columnspan = 1)

extract_all_fields_to_json = Button(main_window_of_gui, text ="All fields json", width = 15, height = 3, command = get_all_fields_to_json)
extract_all_fields_to_json.grid(row = 1, column = 3, rowspan = 3, columnspan = 1)

button_open_the_folder = Button(main_window_of_gui, text = "Open the folder", width = 15, height = 2, command = open_apps_folder)
button_open_the_folder.grid(row = 4, column = 2, rowspan = 2, columnspan = 1)

web_page_link = Label(main_window_of_gui, text = "Documentation", fg = "blue", cursor = "hand2")
web_page_link.grid(row = 1, column = 1, columnspan = 1)
web_page_link.bind("<Button-1>", open_web_link)

text_box = Listbox(main_window_of_gui, height=8)
text_box.grid(column=0, row=12, columnspan=6, sticky=(N,W,E,S))
main_window_of_gui.grid_columnconfigure(0, weight=1)
main_window_of_gui.grid_rowconfigure(12, weight=1)

my_scrollbar = Scrollbar(main_window_of_gui, orient=VERTICAL, command=text_box.yview)
my_scrollbar.grid(column=7, row=12, sticky=(N,S))

text_box['yscrollcommand'] = my_scrollbar.set

main_window_of_gui.mainloop()
